"""
年报和TTM数据生成器
生成过去10年年报 + 最新一期TTM的重构报表
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import logging


class AnnualReportGenerator:
    """年报和TTM数据生成器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def generate_annual_reports_with_ttm(
        self,
        balance_restructured: pd.DataFrame,
        income_restructured: pd.DataFrame,
        cashflow_restructured: pd.DataFrame,
        years: int = 10,
        ts_code: str = None
    ) -> Dict[str, pd.DataFrame]:
        """
        生成年报+TTM的重构报表
        
        Args:
            balance_restructured: 重构后的资产负债表
            income_restructured: 重构后的利润表
            cashflow_restructured: 重构后的现金流量表
            years: 年数，默认10年
            
        Returns:
            包含三大报表的字典
        """
        # 获取最新报告期
        all_dates = self._get_all_dates(balance_restructured, income_restructured, cashflow_restructured)
        if not all_dates:
            self.logger.error("无法获取报告期数据")
            return {}
        
        latest_date = max(all_dates)
        latest_year = int(latest_date[:4])
        latest_month = int(latest_date[4:6])
        
        self.logger.info(f"最新报告期: {latest_date} (年份: {latest_year}, 月份: {latest_month})")
        
        # 判断最新季度
        latest_quarter = self._get_quarter(latest_date)
        self.logger.info(f"最新季度: Q{latest_quarter}")
        
        # 确定年度范围
        # 如果最新是Q4，则年度范围是 (latest_year-years) 到 latest_year
        # 如果最新是Q1-Q3，则年度范围是 (latest_year-years) 到 (latest_year-1)
        if latest_quarter == 4:
            start_year = latest_year - years
            end_year = latest_year
            
            # 检查利润表数据是否完整，给出警告信息
            income_dates = [col for col in income_restructured.columns if col != '项目']
            if latest_date in income_dates:
                income_data = income_restructured[latest_date]
                has_income_data = income_data.notna().any()
                if not has_income_data:
                    self.logger.warning(f"⚠️  {latest_year} 年报利润表数据尚未披露，相关图表将显示空值")
            else:
                self.logger.warning(f"⚠️  {latest_year} 年报利润表数据尚未披露，相关图表将显示空值")
        else:
            start_year = latest_year - years
            end_year = latest_year - 1
        
        self.logger.info(f"年报年度范围: {start_year} - {end_year}")
        
        # 提取年报数据（按降序排列：从新到旧）
        annual_dates = [f"{year}1231" for year in range(end_year, start_year - 1, -1)]
        
        # 生成资产负债表年报+TTM
        balance_annual_ttm = self._generate_balance_sheet_annual_with_ttm(
            balance_restructured, annual_dates, latest_date, latest_quarter
        )
        
        # 生成利润表年报+TTM
        income_annual_ttm = self._generate_income_statement_annual_with_ttm(
            income_restructured, annual_dates, latest_date, latest_quarter
        )
        
        # 生成现金流量表年报+TTM
        cashflow_annual_ttm = self._generate_cashflow_statement_annual_with_ttm(
            cashflow_restructured, annual_dates, latest_date, latest_quarter
        )
        
        # 生成财务指标汇总 Excel
        self._generate_summary_excel(
            balance_annual_ttm, income_annual_ttm, cashflow_annual_ttm, ts_code
        )
        
        return {
            'balance_sheet': balance_annual_ttm,
            'income_statement': income_annual_ttm,
            'cashflow_statement': cashflow_annual_ttm
        }
    
    def _get_all_dates(self, *dfs: pd.DataFrame) -> List[str]:
        """获取所有报表中的日期"""
        all_dates = set()
        
        for df in dfs:
            if df is not None and len(df) > 0:
                # 获取非"项目"列，并转换为字符串
                date_cols = [str(col) for col in df.columns if col != '项目']
                all_dates.update(date_cols)
        
        return sorted(list(all_dates))
    
    def _get_quarter(self, date_str: str) -> int:
        """根据日期字符串判断季度"""
        month = int(date_str[4:6])
        if month == 3:
            return 1
        elif month == 6:
            return 2
        elif month == 9:
            return 3
        elif month == 12:
            return 4
        else:
            return 0
    
    def _generate_balance_sheet_annual_with_ttm(
        self,
        df: pd.DataFrame,
        annual_dates: List[str],
        latest_date: str,
        latest_quarter: int
    ) -> pd.DataFrame:
        """
        生成资产负债表年报+TTM数据
        
        资产负债表是时点数据，不需要计算TTM
        如果最新是Q4，直接使用年报数据
        如果最新是Q1-Q3，使用最新季度数据作为TTM
        """
        if df is None or len(df) == 0:
            return pd.DataFrame()
        
        self.logger.info("生成资产负债表年报+TTM数据...")
        
        # 确保所有列名都是字符串类型（避免整数列名导致的匹配失败）
        df = df.copy()
        df.columns = [str(col) for col in df.columns]
        
        # 构建输出的列顺序
        output_columns = ['项目']
        
        # 如果最新是Q4，年报数据中已包含最新数据
        if latest_quarter == 4:
            # 只保留年报数据
            output_columns.extend(annual_dates)
            self.logger.info(f"最新为Q4，使用年报数据: {latest_date}")
        else:
            # 添加TTM列（使用最新季度数据）
            ttm_label = f"{latest_date[:4]}Q{latest_quarter}-TTM"
            output_columns.append(ttm_label)
            output_columns.extend(annual_dates)
            self.logger.info(f"最新为Q{latest_quarter}，添加TTM列: {ttm_label}")
        
        # 构建结果DataFrame
        result_data = {}
        
        # 获取所有可用的日期列（按降序排列，转换为字符串）
        all_date_cols = sorted([str(col) for col in df.columns if col != '项目'], reverse=True)
        
        for _, row in df.iterrows():
            item_name = row['项目']
            result_data[item_name] = {}
            
            # 添加TTM数据（如果不是Q4）
            if latest_quarter != 4:
                # 直接使用最新季度的数据，不进行向前填充
                # 资产负债表是时点数据，如果某项目为NaN说明该项目已不再使用（如预收款项→合同负债）
                value = row.get(latest_date, np.nan) if latest_date in df.columns else np.nan
                result_data[item_name][ttm_label] = value
            
            # 添加年报数据
            for date in annual_dates:
                if date in df.columns:
                    result_data[item_name][date] = row.get(date, np.nan)
                else:
                    result_data[item_name][date] = np.nan
        
        # 转换为DataFrame
        df_result = pd.DataFrame(result_data).T
        df_result = df_result.reset_index()
        df_result.columns = output_columns
        
        return df_result
    
    def _generate_income_statement_annual_with_ttm(
        self,
        df: pd.DataFrame,
        annual_dates: List[str],
        latest_date: str,
        latest_quarter: int
    ) -> pd.DataFrame:
        """
        生成利润表年报+TTM数据
        
        利润表是期间数据（累计值），需要计算TTM
        TTM = 今年Q累计 - 去年同Q累计 + 去年Q4累计
        """
        if df is None or len(df) == 0:
            return pd.DataFrame()
        
        self.logger.info("生成利润表年报+TTM数据...")
        
        # 确保所有列名都是字符串类型
        df = df.copy()
        df.columns = [str(col) for col in df.columns]
        
        # 构建输出的列顺序
        output_columns = ['项目']
        
        # 如果最新是Q4，年报数据中已包含最新数据
        if latest_quarter == 4:
            output_columns.extend(annual_dates)
            self.logger.info(f"最新为Q4，使用年报数据: {latest_date}")
        else:
            # 添加TTM列
            ttm_label = f"{latest_date[:4]}Q{latest_quarter}-TTM"
            output_columns.append(ttm_label)
            output_columns.extend(annual_dates)
            self.logger.info(f"最新为Q{latest_quarter}，计算TTM: {ttm_label}")
        
        # 计算TTM需要的日期
        if latest_quarter != 4:
            latest_year = int(latest_date[:4])
            last_year = latest_year - 1
            
            # 今年Q累计
            current_q_date = latest_date
            
            # 去年同Q累计
            last_year_same_q_date = f"{last_year}{latest_date[4:]}"
            
            # 去年Q4累计
            last_year_q4_date = f"{last_year}1231"
            
            self.logger.info(f"TTM计算日期: 当前={current_q_date}, 去年同期={last_year_same_q_date}, 去年Q4={last_year_q4_date}")
        
        # 构建结果DataFrame
        result_data = {}
        
        for _, row in df.iterrows():
            item_name = row['项目']
            result_data[item_name] = {}
            
            # 计算TTM（如果不是Q4）
            if latest_quarter != 4:
                ttm_value = self._calculate_ttm(
                    row, current_q_date, last_year_same_q_date, last_year_q4_date, item_name
                )
                result_data[item_name][ttm_label] = ttm_value
            
            # 添加年报数据
            for date in annual_dates:
                if date in df.columns:
                    result_data[item_name][date] = row.get(date, np.nan)
                else:
                    result_data[item_name][date] = np.nan
        
        # 转换为DataFrame
        df_result = pd.DataFrame(result_data).T
        df_result = df_result.reset_index()
        df_result.columns = output_columns
        
        # 重新计算TTM的比率类指标（如果有TTM列）
        if latest_quarter != 4:
            self._recalculate_income_ratios(df_result, ttm_label)
        
        return df_result
    
    def _recalculate_income_ratios(self, df: pd.DataFrame, ttm_col: str):
        """
        重新计算利润表中的比率类指标
        
        比率类指标不能用TTM公式（加减法）计算，必须用对应的分子/分母重新计算
        """
        def get_value(item_name):
            row = df[df['项目'] == item_name]
            if len(row) > 0 and ttm_col in df.columns:
                val = row[ttm_col].values[0]
                return val if pd.notna(val) else 0
            return 0
        
        def set_value(item_name, value):
            df.loc[df['项目'] == item_name, ttm_col] = value
        
        # 获取基础数据
        营业收入 = get_value('营业收入')
        营业成本 = get_value('营业成本')
        毛利 = get_value('毛利')
        税金及附加 = get_value('税金及附加')
        销售费用 = get_value('销售费用')
        管理费用 = get_value('管理费用')
        研发费用 = get_value('研发费用')
        资产减值损失 = get_value('资产减值损失')
        信用减值损失 = get_value('信用减值损失')
        净利润 = get_value('净利润')
        
        # 重新计算比率
        if 营业收入 != 0:
            set_value('营业成本率', 营业成本 / 营业收入)
            set_value('毛利率', 毛利 / 营业收入)
            set_value('营业税金及附加率', 税金及附加 / 营业收入)
            set_value('销售费用率', 销售费用 / 营业收入)
            set_value('管理费用率', 管理费用 / 营业收入)
            set_value('研发费用率', 研发费用 / 营业收入)
            set_value('资产减值损失率', (资产减值损失 + 信用减值损失) / 营业收入)
            set_value('净利润率', 净利润 / 营业收入)
            
            # 其他可能的比率指标
            息税前经营利润 = get_value('息税前经营利润')
            if 息税前经营利润 != 0:
                set_value('息税前经营利润率', 息税前经营利润 / 营业收入)
    
    def _generate_cashflow_statement_annual_with_ttm(
        self,
        df: pd.DataFrame,
        annual_dates: List[str],
        latest_date: str,
        latest_quarter: int
    ) -> pd.DataFrame:
        """
        生成现金流量表年报+TTM数据
        
        现金流量表是期间数据（累计值），需要计算TTM
        TTM = 今年Q累计 - 去年同Q累计 + 去年Q4累计
        """
        if df is None or len(df) == 0:
            return pd.DataFrame()
        
        self.logger.info("生成现金流量表年报+TTM数据...")
        
        # 确保所有列名都是字符串类型
        df = df.copy()
        df.columns = [str(col) for col in df.columns]
        
        # 构建输出的列顺序
        output_columns = ['项目']
        
        # 如果最新是Q4，年报数据中已包含最新数据
        if latest_quarter == 4:
            output_columns.extend(annual_dates)
            self.logger.info(f"最新为Q4，使用年报数据: {latest_date}")
        else:
            # 添加TTM列
            ttm_label = f"{latest_date[:4]}Q{latest_quarter}-TTM"
            output_columns.append(ttm_label)
            output_columns.extend(annual_dates)
            self.logger.info(f"最新为Q{latest_quarter}，计算TTM: {ttm_label}")
        
        # 计算TTM需要的日期
        if latest_quarter != 4:
            latest_year = int(latest_date[:4])
            last_year = latest_year - 1
            
            # 今年Q累计
            current_q_date = latest_date
            
            # 去年同Q累计
            last_year_same_q_date = f"{last_year}{latest_date[4:]}"
            
            # 去年Q4累计
            last_year_q4_date = f"{last_year}1231"
            
            self.logger.info(f"TTM计算日期: 当前={current_q_date}, 去年同期={last_year_same_q_date}, 去年Q4={last_year_q4_date}")
        
        # 构建结果DataFrame
        result_data = {}
        
        for _, row in df.iterrows():
            item_name = row['项目']
            result_data[item_name] = {}
            
            # 计算TTM（如果不是Q4）
            if latest_quarter != 4:
                ttm_value = self._calculate_ttm(
                    row, current_q_date, last_year_same_q_date, last_year_q4_date, item_name
                )
                result_data[item_name][ttm_label] = ttm_value
            
            # 添加年报数据
            for date in annual_dates:
                if date in df.columns:
                    result_data[item_name][date] = row.get(date, np.nan)
                else:
                    result_data[item_name][date] = np.nan
        
        # 转换为DataFrame
        df_result = pd.DataFrame(result_data).T
        df_result = df_result.reset_index()
        df_result.columns = output_columns
        
        # 重新计算TTM的比率类指标（如果有TTM列）
        if latest_quarter != 4:
            self._recalculate_cashflow_ratios(df_result, ttm_label)
        
        return df_result
    
    def _recalculate_cashflow_ratios(self, df: pd.DataFrame, ttm_col: str):
        """
        重新计算现金流量表中的比率类指标
        
        比率类指标不能用TTM公式（加减法）计算，必须用对应的分子/分母重新计算
        """
        def get_value(item_name):
            row = df[df['项目'] == item_name]
            if len(row) > 0 and ttm_col in df.columns:
                val = row[ttm_col].values[0]
                return val if pd.notna(val) else 0
            return 0
        
        def set_value(item_name, value):
            df.loc[df['项目'] == item_name, ttm_col] = value
        
        # 获取基础数据
        销售收到现金 = get_value('销售商品、提供劳务收到的现金')
        营业收入 = get_value('营业收入')
        购买支付现金 = get_value('购买商品、接收劳务支付的现金')
        职工支付现金 = get_value('支付给职工及为职工支付的现金')
        营业总成本 = get_value('营业总成本')
        经营现金流 = get_value('经营活动产生的现金流量净额')
        息前税后经营利润 = get_value('息前税后经营利润')
        净利润 = get_value('净利润')
        
        # 重新计算比率
        if 营业收入 != 0:
            set_value('口径一收入现金含量', 销售收到现金 / 营业收入)
        
        if 营业总成本 != 0:
            set_value('成本费用付现率', (购买支付现金 / 1.17 + 职工支付现金) / 营业总成本)
        
        if 息前税后经营利润 != 0:
            set_value('息前税后经营利润现金含量', 经营现金流 / 息前税后经营利润)
        
        if 净利润 != 0:
            set_value('净利润现金含量', 经营现金流 / 净利润)
    
    def _calculate_ttm(
        self,
        row: pd.Series,
        current_q_date: str,
        last_year_same_q_date: str,
        last_year_q4_date: str,
        item_name: str
    ) -> float:
        """
        计算TTM值
        
        TTM = 今年Q累计 - 去年同Q累计 + 去年Q4累计
        
        Args:
            row: 数据行
            current_q_date: 今年Q累计日期
            last_year_same_q_date: 去年同Q累计日期
            last_year_q4_date: 去年Q4累计日期
            item_name: 项目名称
            
        Returns:
            TTM值
        """
        # 获取值
        current_value = row.get(current_q_date, np.nan)
        last_year_same_q_value = row.get(last_year_same_q_date, np.nan)
        last_year_q4_value = row.get(last_year_q4_date, np.nan)
        
        # 检查是否有足够的有效值
        if pd.isna(current_value) or pd.isna(last_year_q4_value):
            return np.nan
        
        # 如果去年同Q数据为空，说明是Q1数据
        # Q1的TTM = Q1累计 - 0 + 去年Q4累计
        if pd.isna(last_year_same_q_value):
            last_year_same_q_value = 0
        
        # 计算TTM
        ttm = current_value - last_year_same_q_value + last_year_q4_value
        
        return ttm
    
    def format_annual_report(self, df: pd.DataFrame, statement_type: str) -> pd.DataFrame:
        """
        格式化年报报表，使其更易读
        
        Args:
            df: 年报数据
            statement_type: 报表类型（balance_sheet, income_statement, cashflow_statement）
            
        Returns:
            格式化后的DataFrame
        """
        if df is None or len(df) == 0:
            return df
        
        # 格式化日期列名
        df_formatted = df.copy()
        
        # 重命名日期列
        new_columns = {}
        for col in df_formatted.columns:
            if col == '项目':
                new_columns[col] = col
            elif col.endswith('-TTM'):
                new_columns[col] = col
            else:
                # 将 YYYYMMDD 转换为 YYYY/MM/DD
                if len(col) == 8 and col.isdigit():
                    formatted_date = f"{col[:4]}/{col[4:6]}/{col[6:8]}"
                    new_columns[col] = formatted_date
                else:
                    new_columns[col] = col
        
        df_formatted = df_formatted.rename(columns=new_columns)
        
        return df_formatted
    
    def _generate_summary_excel_from_charts(
        self,
        balance_data: pd.DataFrame,
        income_data: pd.DataFrame,
        cashflow_data: pd.DataFrame,
        charts_config: list,
        date_columns: list,
        ts_code: str = None
    ):
        """
        从 HTML 报告生成器返回的图表数据生成财务指标汇总 Excel 文件
        
        Args:
            balance_data: 资产负债表年报+TTM数据
            income_data: 利润表年报+TTM数据
            cashflow_data: 现金流量表年报+TTM数据
            charts_config: HTML 报告生成器返回的图表配置列表
            date_columns: 日期列（从远到近）
            ts_code: 股票代码
        """
        try:
            # 日期列需要反转（从新到旧）
            date_columns_reversed = list(reversed(date_columns))
            
            # 创建汇总数据字典
            summary_data = {'指标': []}
            
            # 添加日期列
            for date_col in date_columns_reversed:
                summary_data[date_col] = []
            
            # 添加关键财务指标（保留原有的）
            summary_data['指标'].append('')
            for date_col in date_columns_reversed:
                summary_data[date_col].append(None)
            
            summary_data['指标'].append('【关键财务指标】（单位：元、%）')
            for date_col in date_columns_reversed:
                summary_data[date_col].append(None)
            
            # 关键指标数据
            key_indicators = [
                ('营业收入', income_data, '营业收入'),
                ('息税前经营利润', income_data, '息税前经营利润'),
                ('(其中)利息费用', income_data, '(其中)利息费用'),
                ('所有者权益合计', balance_data, '所有者权益合计'),
                ('有息债务合计', balance_data, '有息债务合计'),
                ('金融资产合计', balance_data, '金融资产合计'),
                ('长期股权投资', balance_data, '长期股权投资'),
                ('少数股东权益', balance_data, '少数股东权益'),
                ('总股本', balance_data, '期末总股本'),
                ('实际所得税税率', income_data, '实际所得税税率'),
            ]
            
            # 百分比字段列表（值为小数0-1，与income_statement_restructure保持一致）
            _pct_fields = {
                '实际所得税税率', '营业成本率', '毛利率', '净利润率', '息税前经营利润率',
                '销售费用率', '管理费用率', '研发费用率', '营业税金及附加率', '资产减值损失率',
                '有息债务率', '财务成本负担率', '扩张性资本支出占长期资产期初净额的比例',
                '长期股权投资收益率', 'ROIC', 'ROE', '营业外收支及其他占营业收入的比例'
            }
            
            for display_name, df, field_name in key_indicators:
                summary_data['指标'].append(display_name)
                field_row = df[df['项目'] == field_name]
                if len(field_row) > 0:
                    for date_col in date_columns_reversed:
                        if date_col in df.columns:
                            value = field_row[date_col].values[0]
                            if pd.notna(value):
                                # 保持小数形式(0-1)，与income_statement_restructure一致
                                summary_data[date_col].append(round(float(value), 4))
                            else:
                                summary_data[date_col].append(None)
                        else:
                            summary_data[date_col].append(None)
                else:
                    for date_col in date_columns_reversed:
                        summary_data[date_col].append(None)
            
            # marginal sales/capital ratio
            summary_data['指标'].append('marginal sales/capital ratio')
            revenue_row = income_data[income_data['项目'] == '营业收入']
            equity_row = balance_data[balance_data['项目'] == '所有者权益合计']
            debt_row = balance_data[balance_data['项目'] == '有息债务合计']
            financial_row = balance_data[balance_data['项目'] == '金融资产合计']
            
            for i, date_col in enumerate(date_columns_reversed):
                if i == len(date_columns_reversed) - 1:
                    summary_data[date_col].append(None)
                else:
                    prev_date_col = date_columns_reversed[i + 1]
                    if (len(revenue_row) > 0 and len(equity_row) > 0 and len(debt_row) > 0 and len(financial_row) > 0):
                        curr_revenue = revenue_row[date_col].values[0] if date_col in income_data.columns else None
                        prev_revenue = revenue_row[prev_date_col].values[0] if prev_date_col in income_data.columns else None
                        curr_equity = equity_row[date_col].values[0] if date_col in balance_data.columns else None
                        curr_debt = debt_row[date_col].values[0] if date_col in balance_data.columns else None
                        curr_financial = financial_row[date_col].values[0] if date_col in balance_data.columns else None
                        prev_equity = equity_row[prev_date_col].values[0] if prev_date_col in balance_data.columns else None
                        prev_debt = debt_row[prev_date_col].values[0] if prev_date_col in balance_data.columns else None
                        prev_financial = financial_row[prev_date_col].values[0] if prev_date_col in balance_data.columns else None
                        
                        if all(pd.notna(x) for x in [curr_revenue, prev_revenue, curr_equity, curr_debt, curr_financial, prev_equity, prev_debt, prev_financial]):
                            curr_capital = float(curr_equity) + float(curr_debt) - float(curr_financial)
                            prev_capital = float(prev_equity) + float(prev_debt) - float(prev_financial)
                            delta_revenue = float(curr_revenue) - float(prev_revenue)
                            delta_capital = curr_capital - prev_capital
                            if delta_capital != 0:
                                summary_data[date_col].append(round(delta_revenue / delta_capital, 2))
                            else:
                                summary_data[date_col].append(None)
                        else:
                            summary_data[date_col].append(None)
                    else:
                        summary_data[date_col].append(None)
            
            # sales/capital ratio
            summary_data['指标'].append('sales/capital ratio')
            for date_col in date_columns_reversed:
                if (len(revenue_row) > 0 and len(equity_row) > 0 and len(debt_row) > 0 and len(financial_row) > 0):
                    revenue = revenue_row[date_col].values[0] if date_col in income_data.columns else None
                    equity = equity_row[date_col].values[0] if date_col in balance_data.columns else None
                    debt = debt_row[date_col].values[0] if date_col in balance_data.columns else None
                    financial = financial_row[date_col].values[0] if date_col in balance_data.columns else None
                    
                    if all(pd.notna(x) for x in [revenue, equity, debt, financial]):
                        capital = float(equity) + float(debt) - float(financial)
                        if capital != 0:
                            summary_data[date_col].append(round(float(revenue) / capital, 2))
                        else:
                            summary_data[date_col].append(None)
                    else:
                        summary_data[date_col].append(None)
                else:
                    summary_data[date_col].append(None)
            
            # 从图表数据中提取所有指标
            for chart in charts_config:
                # 添加空行
                summary_data['指标'].append('')
                for date_col in date_columns_reversed:
                    summary_data[date_col].append(None)
                
                # 添加图表标题
                chart_title = f"【{chart['title']}】（单位：亿元、%）"
                summary_data['指标'].append(chart_title)
                for date_col in date_columns_reversed:
                    summary_data[date_col].append(None)
                
                # 添加图表中的每个数据系列
                series = chart['data'].get('series', {})
                for series_name, series_data in series.items():
                    summary_data['指标'].append(series_name)
                    data_values = series_data.get('data', [])
                    # 注意：图表数据是从远到近，需要反转以匹配 date_columns_reversed（从新到旧）
                    data_values_reversed = list(reversed(data_values))
                    for i, date_col in enumerate(date_columns_reversed):
                        if i < len(data_values_reversed):
                            value = data_values_reversed[i]
                            if value is not None:
                                fval = float(value)
                                # 图表数据中比率字段已被html_report_generator乘以100，需除以100还原为小数
                                if series_name in _pct_fields:
                                    fval = fval / 100
                                summary_data[date_col].append(round(fval, 4))
                            else:
                                summary_data[date_col].append(None)
                        else:
                            summary_data[date_col].append(None)
            
            # 创建DataFrame
            summary_df = pd.DataFrame(summary_data)
            
            # 将空字符串的指标行保留为空字符串（用于分隔不同类别）
            summary_df['指标'] = summary_df['指标'].replace('', '---')
            
            # 重命名日期列，添加前缀避免 Excel 自动格式化为日期
            # 例如：20251231 -> Y20251231
            rename_dict = {}
            for col in summary_df.columns:
                if col != '指标' and isinstance(col, str) and col.isdigit() and len(col) == 8:
                    rename_dict[col] = f"Y{col}"
            if rename_dict:
                summary_df = summary_df.rename(columns=rename_dict)
            
            # 生成Excel文件路径
            import os
            from datetime import datetime
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_filename = f"data/{ts_code}_financial_summary_{timestamp}.xlsx"
            
            # 使用openpyxl直接写入，避免pandas自动格式化日期
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
            
            # 创建新工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # 写入表头（日期列保持原始字符串格式）
            headers = ['指标'] + date_columns_reversed
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                # 日期列设置为文本格式
                if col_idx > 1:
                    cell.number_format = numbers.FORMAT_TEXT
            
            # 写入数据
            for row_idx, (_, row_data) in enumerate(summary_df.iterrows(), 2):
                for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # 日期列的所有单元格都设置为文本格式
                    if col_idx > 1:
                        cell.number_format = numbers.FORMAT_TEXT
            
            # 格式化表头
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 定义边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 格式化数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                indicator_name = row[0].value
                
                # 空行（分隔符）：无边框，白色背景
                if indicator_name == '---':
                    for cell in row:
                        cell.border = Border()
                        cell.fill = PatternFill(fill_type=None)
                    continue
                
                # 标题行：绿色背景，加粗，居中
                if indicator_name and indicator_name.startswith('【') and indicator_name.endswith('】'):
                    title_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    title_font = Font(bold=True, size=11)
                    for cell in row:
                        cell.fill = title_fill
                        cell.font = title_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    continue
                
                # 普通数据行
                # 指标名称列：加粗，左对齐
                row[0].font = Font(bold=True)
                row[0].alignment = Alignment(horizontal='left', vertical='center')
                row[0].border = thin_border
                
                # 数值列：右对齐
                for cell in row[1:]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = thin_border
                    
                    value = cell.value
                    
                    # 根据指标类型格式化数值
                    if value is not None and isinstance(value, (int, float)) and not isinstance(value, str):
                        # 整数字段（不显示小数点）- 金额类
                        integer_fields = [
                            '营业收入', '息税前经营利润', '(其中)利息费用',
                            '所有者权益合计', '有息债务合计', '金融资产合计',
                            '长期股权投资', '少数股东权益', '总股本',
                            '营业成本', '毛利', '净利润', '息前税后经营利润',
                            '销售费用', '管理费用', '研发费用', '税金及附加', '资产减值损失',
                            '经营资产合计', '周转性经营投入合计', '长期经营资产合计',
                            '短期债务', '长期债务', '资本总额', '存货', '应收账款', '应付账款', '货币资金',
                            '经营活动产生的现金流量净额', '投资活动产生的现金流净额', '资本支出总额',
                            '经营资产自由现金流量', '债务筹资净额', '长期经营资产净投资额', '扩张性资本支出',
                            '真实财务费用', '税前利润', '资本支出净额', '长期股权投资收益', '长期股权外投资收益',
                            'FCFE', '分红'
                        ]
                        
                        # 百分比字段（已经是百分比数值，如 15.5 表示 15.5%）
                        percentage_fields = [
                            '实际所得税税率', '营业成本率', '毛利率', '净利润率', '息税前经营利润率',
                            '销售费用率', '管理费用率', '研发费用率', '营业税金及附加率', '资产减值损失率',
                            '有息债务率', '财务成本负担率', '扩张性资本支出占长期资产期初净额的比例',
                            '长期股权投资收益率', 'ROIC', 'ROE'
                        ]
                        
                        # 比率字段（2位小数）
                        ratio_fields = [
                            'marginal sales/capital ratio', 'sales/capital ratio',
                            '经营资产周转率', '长期经营资产周转率', '固定资产周转率',
                            '应收账款周转率', '存货周转率', '应付账款周转率'
                        ]
                        
                        # 天数字段（1位小数）
                        days_fields = [
                            '应收账款周转天数', '存货周转天数', '应付账款周转天数',
                            '营业周期', '现金周期'
                        ]
                        
                        if indicator_name in integer_fields:
                            # 整数格式，带千分位分隔符
                            cell.number_format = '#,##0'
                        elif indicator_name in percentage_fields:
                            # 百分比格式，1位小数（数值为小数0-1，Excel自动乘100显示）
                            cell.number_format = '0.0%'
                        elif indicator_name in ratio_fields:
                            # 比率格式，2位小数
                            cell.number_format = '0.00'
                        elif indicator_name in days_fields:
                            # 天数格式，1位小数
                            cell.number_format = '0.0'
                        else:
                            # 默认格式
                            if abs(value) >= 1000:
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '0.00'
            
            # 调整列宽
            ws.column_dimensions['A'].width = 30  # 指标列
            for col_idx in range(2, len(date_columns_reversed) + 2):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 15
            
            # 冻结首行和首列
            ws.freeze_panes = 'B2'
            
            # 保存
            wb.save(excel_filename)
            wb.close()
            
            print(f"✓ 财务指标汇总已导出到: {excel_filename}")
            
        except Exception as e:
            print(f"生成财务指标汇总Excel失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _generate_summary_excel(
        self,
        balance_data: pd.DataFrame,
        income_data: pd.DataFrame,
        cashflow_data: pd.DataFrame,
        ts_code: str = None
    ):
        """
        生成财务指标汇总 Excel 文件（旧方法，保留用于向后兼容）
        直接使用 HTML 报告生成器中已经计算好的图表数据
        
        Args:
            balance_data: 资产负债表年报+TTM数据
            income_data: 利润表年报+TTM数据
            cashflow_data: 现金流量表年报+TTM数据
            ts_code: 股票代码
        """
        try:
            from html_report_generator import FinancialStatementsReportGenerator
            
            # 获取日期列（降序：从新到旧）
            date_columns = [col for col in balance_data.columns if col != '项目']
            
            # 创建 HTML 报告生成器实例以获取图表数据
            html_generator = FinancialStatementsReportGenerator(
                company_name=self.company_name if hasattr(self, 'company_name') else "",
                stock_code=ts_code if ts_code else ""
            )
            
            # 生成所有图表配置（包含已计算的数据）
            all_charts = []
            
            # 1. 利润分析图表
            profit_charts = html_generator._generate_profit_charts(income_data, date_columns)
            all_charts.extend(profit_charts)
            
            # 2. 资产负债表分析图表
            balance_charts = html_generator._generate_balance_charts(balance_data, date_columns)
            all_charts.extend(balance_charts)
            
            # 3. 经营效率分析图表
            efficiency_charts = html_generator._generate_efficiency_charts(balance_data, income_data, date_columns)
            # 为效率图表添加计算字段
            for chart in efficiency_charts:
                if chart.get('needs_roe_roic_calc'):
                    html_generator._add_roe_roic_calculations(chart, balance_data, income_data, date_columns)
                elif chart.get('needs_asset_efficiency_calc'):
                    html_generator._add_asset_efficiency_calculations(chart, balance_data, income_data, date_columns)
                elif chart.get('needs_turnover_days_calc'):
                    html_generator._add_turnover_days_calculations(chart, balance_data, income_data, date_columns)
                elif chart.get('needs_turnover_ratio_calc'):
                    html_generator._add_turnover_ratio_calculations(chart, balance_data, income_data, date_columns)
            all_charts.extend(efficiency_charts)
            
            # 4. 财务成本分析图表
            finance_cost_charts = html_generator._generate_finance_cost_charts(income_data, date_columns)
            for chart in finance_cost_charts:
                if chart.get('needs_finance_cost_calc'):
                    html_generator._add_finance_cost_calculations(chart, income_data, date_columns)
            all_charts.extend(finance_cost_charts)
            
            # 5. 长期资产投资和并购活动分析图表
            capex_charts = html_generator._generate_capex_charts(balance_data, cashflow_data, date_columns)
            for chart in capex_charts:
                if chart.get('needs_capex_calc'):
                    html_generator._add_capex_calculations(chart, balance_data, cashflow_data, date_columns)
            all_charts.extend(capex_charts)
            
            # 6. 投资收益分析图表
            investment_charts = html_generator._generate_investment_income_charts(balance_data, income_data, date_columns)
            for chart in investment_charts:
                if chart.get('needs_investment_income_calc'):
                    html_generator._add_investment_income_calculations(chart, balance_data, income_data, date_columns)
            all_charts.extend(investment_charts)
            
            # 7. 现金流分析图表
            cashflow_charts = html_generator._generate_cashflow_charts(cashflow_data, date_columns)
            all_charts.extend(cashflow_charts)
            
            # 8. FCFE vs Dividend 图表
            fcfe_chart = html_generator._generate_fcfe_dividend_chart(balance_data, income_data, cashflow_data, date_columns)
            all_charts.append(fcfe_chart)
            
            # 创建汇总数据字典
            summary_data = {'指标': []}
            
            # 添加日期列
            for date_col in date_columns:
                summary_data[date_col] = []
            
            # 添加关键财务指标（保留原有的）
            summary_data['指标'].append('')
            for date_col in date_columns:
                summary_data[date_col].append(None)
            
            summary_data['指标'].append('【关键财务指标】')
            for date_col in date_columns:
                summary_data[date_col].append(None)
            
            # 关键指标数据
            key_indicators = [
                ('营业收入', income_data, '营业收入'),
                ('息税前经营利润', income_data, '息税前经营利润'),
                ('(其中)利息费用', income_data, '(其中)利息费用'),
                ('所有者权益合计', balance_data, '所有者权益合计'),
                ('有息债务合计', balance_data, '有息债务合计'),
                ('金融资产合计', balance_data, '金融资产合计'),
                ('长期股权投资', balance_data, '长期股权投资'),
                ('少数股东权益', balance_data, '少数股东权益'),
                ('总股本', balance_data, '期末总股本'),
                ('实际所得税税率', income_data, '实际所得税税率'),
            ]
            
            # 百分比字段列表（值为小数0-1，需要乘以100后存储）
            _pct_fields = {
                '实际所得税税率', '营业成本率', '毛利率', '净利润率', '息税前经营利润率',
                '销售费用率', '管理费用率', '研发费用率', '营业税金及附加率', '资产减值损失率',
                '有息债务率', '财务成本负担率', '扩张性资本支出占长期资产期初净额的比例',
                '长期股权投资收益率', 'ROIC', 'ROE', '营业外收支及其他占营业收入的比例'
            }
            
            for display_name, df, field_name in key_indicators:
                summary_data['指标'].append(display_name)
                field_row = df[df['项目'] == field_name]
                if len(field_row) > 0:
                    for date_col in date_columns:
                        if date_col in df.columns:
                            value = field_row[date_col].values[0]
                            if pd.notna(value):
                                # 保持小数形式(0-1)，与income_statement_restructure一致
                                summary_data[date_col].append(round(float(value), 4))
                            else:
                                summary_data[date_col].append(None)
                        else:
                            summary_data[date_col].append(None)
                else:
                    for date_col in date_columns:
                        summary_data[date_col].append(None)
            
            # marginal sales/capital ratio
            summary_data['指标'].append('marginal sales/capital ratio')
            revenue_row = income_data[income_data['项目'] == '营业收入']
            equity_row = balance_data[balance_data['项目'] == '所有者权益合计']
            debt_row = balance_data[balance_data['项目'] == '有息债务合计']
            financial_row = balance_data[balance_data['项目'] == '金融资产合计']
            
            for i, date_col in enumerate(date_columns):
                if i == len(date_columns) - 1:
                    summary_data[date_col].append(None)
                else:
                    prev_date_col = date_columns[i + 1]
                    if (len(revenue_row) > 0 and len(equity_row) > 0 and len(debt_row) > 0 and len(financial_row) > 0):
                        curr_revenue = revenue_row[date_col].values[0] if date_col in income_data.columns else None
                        prev_revenue = revenue_row[prev_date_col].values[0] if prev_date_col in income_data.columns else None
                        curr_equity = equity_row[date_col].values[0] if date_col in balance_data.columns else None
                        curr_debt = debt_row[date_col].values[0] if date_col in balance_data.columns else None
                        curr_financial = financial_row[date_col].values[0] if date_col in balance_data.columns else None
                        prev_equity = equity_row[prev_date_col].values[0] if prev_date_col in balance_data.columns else None
                        prev_debt = debt_row[prev_date_col].values[0] if prev_date_col in balance_data.columns else None
                        prev_financial = financial_row[prev_date_col].values[0] if prev_date_col in balance_data.columns else None
                        
                        if all(pd.notna(x) for x in [curr_revenue, prev_revenue, curr_equity, curr_debt, curr_financial, prev_equity, prev_debt, prev_financial]):
                            curr_capital = float(curr_equity) + float(curr_debt) - float(curr_financial)
                            prev_capital = float(prev_equity) + float(prev_debt) - float(prev_financial)
                            delta_revenue = float(curr_revenue) - float(prev_revenue)
                            delta_capital = curr_capital - prev_capital
                            if delta_capital != 0:
                                summary_data[date_col].append(round(delta_revenue / delta_capital, 2))
                            else:
                                summary_data[date_col].append(None)
                        else:
                            summary_data[date_col].append(None)
                    else:
                        summary_data[date_col].append(None)
            
            # sales/capital ratio
            summary_data['指标'].append('sales/capital ratio')
            for date_col in date_columns:
                if (len(revenue_row) > 0 and len(equity_row) > 0 and len(debt_row) > 0 and len(financial_row) > 0):
                    revenue = revenue_row[date_col].values[0] if date_col in income_data.columns else None
                    equity = equity_row[date_col].values[0] if date_col in balance_data.columns else None
                    debt = debt_row[date_col].values[0] if date_col in balance_data.columns else None
                    financial = financial_row[date_col].values[0] if date_col in balance_data.columns else None
                    
                    if all(pd.notna(x) for x in [revenue, equity, debt, financial]):
                        capital = float(equity) + float(debt) - float(financial)
                        if capital != 0:
                            summary_data[date_col].append(round(float(revenue) / capital, 2))
                        else:
                            summary_data[date_col].append(None)
                    else:
                        summary_data[date_col].append(None)
                else:
                    summary_data[date_col].append(None)
            
            # 从图表数据中提取所有指标
            for chart in all_charts:
                # 添加空行
                summary_data['指标'].append('')
                for date_col in date_columns:
                    summary_data[date_col].append(None)
                
                # 添加图表标题
                chart_title = f"【{chart['title']}】"
                summary_data['指标'].append(chart_title)
                for date_col in date_columns:
                    summary_data[date_col].append(None)
                
                # 添加图表中的每个数据系列
                series = chart['data'].get('series', {})
                for series_name, series_data in series.items():
                    summary_data['指标'].append(series_name)
                    data_values = series_data.get('data', [])
                    for i, date_col in enumerate(date_columns):
                        if i < len(data_values):
                            value = data_values[i]
                            if value is not None:
                                fval = float(value)
                                # 图表数据中比率字段已被html_report_generator乘以100，需除以100还原为小数
                                if series_name in _pct_fields:
                                    fval = fval / 100
                                summary_data[date_col].append(round(fval, 4))
                            else:
                                summary_data[date_col].append(None)
                        else:
                            summary_data[date_col].append(None)
            
            
            # 创建DataFrame
            summary_df = pd.DataFrame(summary_data)
            
            # 将空字符串的指标行保留为空字符串（用于分隔不同类别）
            summary_df['指标'] = summary_df['指标'].replace('', '---')
            
            # 生成Excel文件路径
            import os
            from datetime import datetime
            
            # 生成时间戳
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # 使用传入的股票代码或默认文件名
            if ts_code:
                excel_path = f'data/{ts_code}_financial_summary_{timestamp}.xlsx'
            else:
                excel_path = f'data/financial_summary_{timestamp}.xlsx'
            
            # 导出到Excel并美化格式
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "财务指标汇总"
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 设置边框
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
                    
                    # 表头样式
                    if r_idx == 1:
                        cell.font = Font(bold=True, size=11, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        # 获取当前行的指标名称
                        indicator_name = ws.cell(r_idx, 1).value
                        
                        # 图表标题行样式（以【】包围的）
                        if indicator_name and indicator_name.startswith('【') and indicator_name.endswith('】'):
                            if c_idx == 1:
                                cell.font = Font(bold=True, size=12, color="FFFFFF")
                                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            else:
                                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 空行样式（指标名为 '---' 的）
                        elif indicator_name == '---':
                            # 空行不显示边框
                            cell.border = Border()
                            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        
                        # 数据行样式
                        else:
                            if c_idx == 1:
                                # 指标列左对齐
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                                cell.font = Font(bold=True, size=10)
                            else:
                                # 数值列右对齐
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                                
                                # 根据指标类型格式化数值
                                if value is not None and isinstance(value, (int, float)) and not isinstance(value, str):
                                    # 整数字段（不显示小数点）- 金额类
                                    integer_fields = [
                                        '营业收入', '息税前经营利润', '(其中)利息费用',
                                        '所有者权益合计', '有息债务合计', '金融资产合计',
                                        '长期股权投资', '少数股东权益', '总股本',
                                        '营业成本', '毛利', '净利润', '息前税后经营利润',
                                        '销售费用', '管理费用', '研发费用', '税金及附加', '资产减值损失',
                                        '经营资产合计', '周转性经营投入合计', '长期经营资产合计',
                                        '短期债务', '长期债务', '资本总额', '存货', '应收账款', '应付账款', '货币资金',
                                        '经营活动产生的现金流量净额', '投资活动产生的现金流净额', '资本支出总额',
                                        '经营资产自由现金流量', '债务筹资净额', '长期经营资产净投资额', '扩张性资本支出',
                                        '真实财务费用', '税前利润', '资本支出净额', '长期股权投资收益', '长期股权外投资收益',
                                        'FCFE', '分红'
                                    ]
                                    
                                    # 百分比字段（已经是百分比数值，如 15.5 表示 15.5%）
                                    percentage_fields = [
                                        '实际所得税税率', '营业成本率', '毛利率', '净利润率', '息税前经营利润率',
                                        '销售费用率', '管理费用率', '研发费用率', '营业税金及附加率', '资产减值损失率',
                                        '有息债务率', '财务成本负担率', '扩张性资本支出占长期资产期初净额的比例',
                                        '长期股权投资收益率', 'ROIC', 'ROE'
                                    ]
                                    
                                    # 比率字段（2位小数）
                                    ratio_fields = [
                                        'marginal sales/capital ratio', 'sales/capital ratio',
                                        '经营资产周转率', '长期经营资产周转率', '固定资产周转率',
                                        '应收账款周转率', '存货周转率', '应付账款周转率'
                                    ]
                                    
                                    # 天数字段（1位小数）
                                    days_fields = [
                                        '应收账款周转天数', '存货周转天数', '应付账款周转天数',
                                        '营业周期', '现金周期'
                                    ]
                                    
                                    if indicator_name in integer_fields:
                                        # 整数格式，带千分位分隔符
                                        cell.number_format = '#,##0'
                                    elif indicator_name in percentage_fields:
                                        # 百分比格式，1位小数（数值为小数0-1，Excel自动乘100显示）
                                        cell.number_format = '0.0%'
                                    elif indicator_name in ratio_fields:
                                        # 比率格式，2位小数
                                        cell.number_format = '0.00'
                                    elif indicator_name in days_fields:
                                        # 天数格式，1位小数
                                        cell.number_format = '0.0'
                                    else:
                                        # 默认格式
                                        if abs(value) >= 1000:
                                            cell.number_format = '#,##0.00'
                                        else:
                                            cell.number_format = '0.00'
            
            # 调整列宽
            ws.column_dimensions['A'].width = 30  # 指标列
            for col_idx in range(2, len(date_columns) + 2):
                ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 15
            
            # 冻结首行和首列
            ws.freeze_panes = 'B2'
            
            # 保存文件
            wb.save(excel_path)
            
            print(f"✓ 财务指标汇总已导出到: {excel_path}")
            
        except Exception as e:
            self.logger.error(f"生成财务指标汇总Excel失败: {e}")
            import traceback
            traceback.print_exc()


# ============================================================================
# 主函数测试
# ============================================================================

if __name__ == '__main__':
    import sys
    import os
    
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    # 测试数据路径
    data_dir = 'data'
    ts_code = '603345.SH'
    
    # 读取重构后的数据
    balance_file = os.path.join(data_dir, f'{ts_code}_balancesheet_restructured.csv')
    income_file = os.path.join(data_dir, f'{ts_code}_income_restructured.csv')
    cashflow_file = os.path.join(data_dir, f'{ts_code}_cashflow_restructured.csv')
    
    print("读取重构后的数据...")
    df_balance = pd.read_csv(balance_file, encoding='utf-8-sig')
    df_income = pd.read_csv(income_file, encoding='utf-8-sig')
    df_cashflow = pd.read_csv(cashflow_file, encoding='utf-8-sig')
    
    print(f"资产负债表: {df_balance.shape}")
    print(f"利润表: {df_income.shape}")
    print(f"现金流量表: {df_cashflow.shape}")
    
    # 生成年报+TTM数据
    generator = AnnualReportGenerator()
    reports = generator.generate_annual_reports_with_ttm(
        df_balance, df_income, df_cashflow, years=10
    )
    
    # 格式化并保存
    for report_name, df_report in reports.items():
        if df_report is not None and len(df_report) > 0:
            # 格式化
            df_formatted = generator.format_annual_report(df_report, report_name)
            
            # 保存
            output_file = os.path.join(data_dir, f'{ts_code}_{report_name}_annual_ttm.csv')
            df_formatted.to_csv(output_file, index=False, encoding='utf-8-sig')
            print(f"\n✓ {report_name} 已保存到: {output_file}")
            
            # 显示前几行
            print(f"\n{report_name} 前5行:")
            print(df_formatted.head().to_string())
