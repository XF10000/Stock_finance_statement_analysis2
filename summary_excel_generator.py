#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
简要经营数据汇总Excel报告生成器
"""

import pandas as pd
import tushare as ts
from typing import Dict, List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class SummaryExcelGenerator:
    """简要经营数据汇总Excel报告生成器"""
    
    def __init__(self, ts_token: str):
        """
        初始化生成器
        
        Args:
            ts_token: Tushare API token
        """
        self.ts_token = ts_token
        ts.set_token(ts_token)
        self.pro = ts.pro_api()
    
    def generate_report(
        self,
        balance_annual_df: pd.DataFrame,
        income_annual_df: pd.DataFrame,
        cashflow_annual_df: pd.DataFrame,
        balance_ttm_df: pd.DataFrame,
        income_ttm_df: pd.DataFrame,
        cashflow_ttm_df: pd.DataFrame,
        stock_code: str,
        output_path: str
    ) -> str:
        """
        生成简要经营数据汇总Excel报告
        
        Args:
            balance_annual_df: 年报资产负债表
            income_annual_df: 年报利润表
            cashflow_annual_df: 年报现金流量表
            balance_ttm_df: TTM资产负债表
            income_ttm_df: TTM利润表
            cashflow_ttm_df: TTM现金流量表
            stock_code: 股票代码
            output_path: 输出文件路径
            
        Returns:
            生成的Excel文件路径
        """
        # 获取日期列
        date_columns = [col for col in balance_annual_df.columns if col != '项目']
        
        # 提取年报数据
        annual_data = self._extract_data(
            balance_annual_df, income_annual_df, cashflow_annual_df, date_columns
        )
        
        # 提取TTM数据
        ttm_data = self._extract_data(
            balance_ttm_df, income_ttm_df, cashflow_ttm_df, date_columns
        )
        
        # 从资产负债表读取总股本数据
        total_share_data = self._get_total_share_from_balance(balance_annual_df, date_columns)
        annual_data['总股本'] = total_share_data
        ttm_data['总股本'] = total_share_data
        
        # 从分红送股.xlsx读取分红数据
        dividend_data = self._get_dividend_from_file(stock_code, date_columns, total_share_data)
        
        # 创建Excel文件
        self._create_excel(annual_data, ttm_data, dividend_data, date_columns, output_path)
        
        print(f"✓ 简要经营数据汇总已生成: {output_path}")
        return output_path
    
    def _extract_data(
        self,
        balance_df: pd.DataFrame,
        income_df: pd.DataFrame,
        cashflow_df: pd.DataFrame,
        date_columns: List[str]
    ) -> Dict[str, List]:
        """提取所需数据"""
        
        data = {}
        
        # 第一部分字段
        data['营业收入'] = self._get_field_data(income_df, '营业收入', date_columns)
        data['息税前经营利润'] = self._get_field_data(income_df, '息税前经营利润', date_columns)
        data['利息费用'] = self._get_field_data(income_df, '(其中)利息费用', date_columns)
        data['所有者权益合计'] = self._get_field_data(balance_df, '所有者权益合计', date_columns)
        data['有息债务合计'] = self._get_field_data(balance_df, '有息债务合计', date_columns)
        data['金融资产合计'] = self._get_field_data(balance_df, '金融资产合计', date_columns)
        data['长期股权投资'] = self._get_field_data(balance_df, '长期股权投资', date_columns)
        data['少数股东权益'] = self._get_field_data(balance_df, '少数股东权益', date_columns)
        # 总股本从Tushare API获取，这里先占位
        data['总股本'] = [None] * len(date_columns)
        data['当前股价'] = [None] * len(date_columns)
        data['实际所得税税率'] = self._get_field_data(income_df, '实际所得税税率', date_columns)
        
        # 第二部分字段
        data['净利润'] = self._get_field_data(income_df, '净利润', date_columns)
        
        # 从现金流量表提取折旧摊销相关
        固定资产折旧 = self._get_field_data(cashflow_df, '固定资产折旧、油气资产折耗、生产性生物资产折旧', date_columns)
        无形资产摊销 = self._get_field_data(cashflow_df, '无形资产摊销', date_columns)
        长期待摊费用摊销 = self._get_field_data(cashflow_df, '长期待摊费用摊销', date_columns)
        处置损失 = self._get_field_data(cashflow_df, '处置固定资产、无形资产和其他长期资产的损失', date_columns)
        固定资产报废损失 = self._get_field_data(cashflow_df, '固定资产报废损失', date_columns)
        
        # 计算折旧及摊销合计
        折旧及摊销合计 = []
        for i in range(len(date_columns)):
            total = 0
            for arr in [固定资产折旧, 无形资产摊销, 长期待摊费用摊销, 处置损失, 固定资产报废损失]:
                if i < len(arr) and arr[i] is not None:
                    total += arr[i]
            折旧及摊销合计.append(total if total != 0 else None)
        data['折旧及摊销合计'] = 折旧及摊销合计
        
        data['资本支出总额'] = self._get_field_data(cashflow_df, '资本支出总额', date_columns)
        data['周转性经营投入合计'] = self._get_field_data(balance_df, '周转性经营投入合计', date_columns)
        
        # 计算营运资本变化量
        周转性经营投入 = data['周转性经营投入合计']
        营运资本变化量 = []
        for i in range(len(date_columns)):
            # 当前期 - 上一期
            if i < len(date_columns) - 1 and 周转性经营投入[i] is not None and 周转性经营投入[i+1] is not None:
                营运资本变化量.append(周转性经营投入[i] - 周转性经营投入[i+1])
            else:
                营运资本变化量.append(None)
        data['营运资本变化量'] = 营运资本变化量
        
        # 计算债务变化
        有息债务 = data['有息债务合计']
        债务变化 = []
        for i in range(len(date_columns)):
            # 当前期 - 上一期
            if i < len(date_columns) - 1 and 有息债务[i] is not None and 有息债务[i+1] is not None:
                债务变化.append(有息债务[i] - 有息债务[i+1])
            else:
                债务变化.append(None)
        data['债务变化'] = 债务变化
        
        # 计算FCFE = 净利润 + 折旧及摊销合计 - 资本支出总额 - 营运资本变化量 + 债务变化
        FCFE = []
        for i in range(len(date_columns)):
            净利润_val = data['净利润'][i] if data['净利润'][i] is not None else 0
            折旧_val = 折旧及摊销合计[i] if 折旧及摊销合计[i] is not None else 0
            资本支出_val = data['资本支出总额'][i] if data['资本支出总额'][i] is not None else 0
            营运资本_val = 营运资本变化量[i] if 营运资本变化量[i] is not None else 0
            债务_val = 债务变化[i] if 债务变化[i] is not None else 0
            
            # 只要有变化量数据就可以计算FCFE
            if 营运资本变化量[i] is not None and 债务变化[i] is not None:
                fcfe_val = 净利润_val + 折旧_val - 资本支出_val - 营运资本_val + 债务_val
                FCFE.append(fcfe_val)
            else:
                FCFE.append(None)
        data['FCFE'] = FCFE
        
        return data
    
    def _get_field_data(self, df: pd.DataFrame, field_name: str, date_columns: List[str]) -> List:
        """从DataFrame中提取指定字段的数据"""
        row = df[df['项目'] == field_name]
        if len(row) == 0:
            return [None] * len(date_columns)
        
        data = []
        for col in date_columns:
            val = row[col].values[0]
            data.append(val if pd.notna(val) else None)
        return data
    
    def _get_total_share_from_balance(self, balance_df: pd.DataFrame, date_columns: List[str]) -> List:
        """从资产负债表读取总股本数据"""
        try:
            # 查找总股本行
            total_share_row = balance_df[balance_df['项目'] == '总股本']
            
            if len(total_share_row) == 0:
                print("警告: 资产负债表中未找到总股本数据")
                return [None] * len(date_columns)
            
            # 提取总股本数据
            total_share_data = []
            for col in date_columns:
                val = total_share_row[col].values[0] if col in total_share_row.columns else None
                total_share_data.append(val if pd.notna(val) else None)
            
            return total_share_data
        except Exception as e:
            print(f"从资产负债表读取总股本失败: {e}")
            return [None] * len(date_columns)
    
    def _get_dividend_from_file(self, stock_code: str, date_columns: List[str], total_share_data: List) -> List:
        """从分红送股.xlsx文件读取分红数据"""
        try:
            import os
            # 读取分红送股.xlsx文件
            dividend_file = f'data/{stock_code}_分红送股.xlsx'
            
            if not os.path.exists(dividend_file):
                print(f"警告: 未找到分红送股文件: {dividend_file}")
                return [None] * len(date_columns)
            
            df = pd.read_excel(dividend_file)
            
            if df is None or len(df) == 0:
                print("警告: 分红送股文件为空")
                return [None] * len(date_columns)
            
            # 处理列名（可能是中文或英文）
            if '报告期' in df.columns:
                df['end_date'] = df['报告期']
            if '每股派息(税后)' in df.columns:
                df['cash_div_tax'] = df['每股派息(税后)']
            if '每股派息(税前)' in df.columns:
                df['cash_div'] = df['每股派息(税前)']
            
            # 将end_date转换为YYYYMMDD格式，并提取年份和月份
            df['end_date'] = pd.to_datetime(df['end_date'], format='%Y%m%d').dt.strftime('%Y%m%d')
            df['year'] = df['end_date'].str[:4]
            df['month'] = df['end_date'].str[4:6]
            
            # 优先使用cash_div_tax，其次cash_div
            df['dividend'] = df.apply(lambda row: row['cash_div_tax'] if pd.notna(row['cash_div_tax']) and row['cash_div_tax'] > 0 else (row['cash_div'] if pd.notna(row['cash_div']) else 0), axis=1)
            
            # 去除重复记录：同一个end_date只保留dividend最大的一条记录
            df = df.sort_values('dividend', ascending=False).drop_duplicates(subset=['end_date'], keep='first')
            
            # 按年份和期间汇总每股派息
            year_dividend_dict = {}  # 全年分红
            year_q1q3_dividend_dict = {}  # Q1-Q3分红（用于TTM）
            
            for _, row in df.iterrows():
                year = row['year']
                month = row['month']
                cash_div = row['dividend']
                
                if cash_div > 0:
                    # 累积全年分红
                    if year not in year_dividend_dict:
                        year_dividend_dict[year] = 0
                    year_dividend_dict[year] += cash_div
                    
                    # 累积Q1-Q3分红（end_date在03/31, 06/30, 09/30的分红）
                    if month in ['03', '06', '09']:
                        if year not in year_q1q3_dividend_dict:
                            year_q1q3_dividend_dict[year] = 0
                        year_q1q3_dividend_dict[year] += cash_div
            
            # 匹配日期列并计算分红总额
            dividend_data = []
            for i, col in enumerate(date_columns):
                # 提取年份，判断是否为TTM数据
                year = col[:4]
                is_ttm = 'TTM' in col or 'Q3' in col
                
                # 对于TTM数据，使用Q1-Q3分红；对于年报数据，使用全年分红
                if is_ttm:
                    cash_div_per_share = year_q1q3_dividend_dict.get(year, 0)
                else:
                    cash_div_per_share = year_dividend_dict.get(year, 0)
                
                # 分红总额 = 每股派息 × 总股本
                if cash_div_per_share > 0 and i < len(total_share_data) and total_share_data[i] is not None:
                    # 每股派息单位是元，总股本单位是股，结果单位是元
                    total_dividend = cash_div_per_share * total_share_data[i]
                    dividend_data.append(total_dividend)
                else:
                    dividend_data.append(None)
            
            return dividend_data
        except Exception as e:
            print(f"从分红送股文件读取数据失败: {e}")
            import traceback
            traceback.print_exc()
            return [None] * len(date_columns)
    
    def _create_excel(
        self,
        annual_data: Dict[str, List],
        ttm_data: Dict[str, List],
        dividend_data: List,
        date_columns: List[str],
        output_path: str
    ):
        """创建Excel文件（横向格式）"""
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "简要经营数据汇总"
        
        # 定义样式
        header_font = Font(name='微软雅黑', size=10, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        normal_font = Font(name='微软雅黑', size=10)
        
        # 第一部分字段
        part1_fields = [
            '营业收入', '息税前经营利润', '利息费用', '所有者权益合计', '有息债务合计',
            '金融资产合计', '长期股权投资', '少数股东权益', '总股本', '当前股价', '实际所得税税率'
        ]
        
        # 第二部分字段
        part2_fields = [
            '净利润', '折旧及摊销合计', '资本支出总额', '周转性经营投入合计',
            '营运资本变化量', '有息债务合计', '债务变化', 'FCFE', '分红'
        ]
        
        row_num = 1
        
        # 写入第一部分（年报数据）
        # 表头行
        ws.cell(row=row_num, column=1, value='年度')
        ws.cell(row=row_num, column=1).font = header_font
        ws.cell(row=row_num, column=1).fill = header_fill
        
        for i, date in enumerate(date_columns):
            # 格式化日期为 YYYY/MM/DD
            formatted_date = f"{date[:4]}/{date[4:6]}/{date[6:]}"
            ws.cell(row=row_num, column=i+2, value=formatted_date)
            ws.cell(row=row_num, column=i+2).font = header_font
            ws.cell(row=row_num, column=i+2).fill = header_fill
            ws.cell(row=row_num, column=i+2).alignment = Alignment(horizontal='right')
        row_num += 1
        
        # 第一部分数据行
        for field in part1_fields:
            ws.cell(row=row_num, column=1, value=field)
            ws.cell(row=row_num, column=1).font = normal_font
            
            data_list = annual_data.get(field, [])
            
            for i, val in enumerate(data_list):
                cell = ws.cell(row=row_num, column=i+2, value=val)
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='right')
                if val is not None and field != '实际所得税税率':
                    cell.number_format = '#,##0'
                elif val is not None and field == '实际所得税税率':
                    cell.number_format = '0.000000000'
            row_num += 1
        
        # 空行
        row_num += 1
        
        # 写入第二部分（TTM数据）
        # 表头行
        ws.cell(row=row_num, column=1, value='年度')
        ws.cell(row=row_num, column=1).font = header_font
        ws.cell(row=row_num, column=1).fill = header_fill
        
        for i, date in enumerate(date_columns):
            formatted_date = f"{date[:4]}/{date[4:6]}/{date[6:]}"
            ws.cell(row=row_num, column=i+2, value=formatted_date)
            ws.cell(row=row_num, column=i+2).font = header_font
            ws.cell(row=row_num, column=i+2).fill = header_fill
            ws.cell(row=row_num, column=i+2).alignment = Alignment(horizontal='right')
        row_num += 1
        
        # 第二部分数据行
        for field in part2_fields:
            ws.cell(row=row_num, column=1, value=field)
            ws.cell(row=row_num, column=1).font = normal_font
            
            if field == '分红':
                data_list = dividend_data
            else:
                data_list = ttm_data.get(field, [])
            
            for i, val in enumerate(data_list):
                cell = ws.cell(row=row_num, column=i+2, value=val)
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='right')
                if val is not None:
                    cell.number_format = '#,##0'
            row_num += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 18
        for i in range(len(date_columns)):
            col_letter = get_column_letter(i + 2)
            ws.column_dimensions[col_letter].width = 16
        
        # 保存文件
        wb.save(output_path)


if __name__ == '__main__':
    import sys
    import yaml
    
    if len(sys.argv) < 2:
        print("用法: python summary_excel_generator.py <股票代码>")
        sys.exit(1)
    
    stock_code = sys.argv[1]
    
    # 读取配置
    with open('config.yaml', 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    
    ts_token = config.get('tushare', {}).get('token')
    if not ts_token:
        print("错误: 请在config.yaml中配置tushare.token")
        sys.exit(1)
    
    # 读取数据
    balance_annual = pd.read_csv(f'data/{stock_code}_balance_sheet_annual_ttm.csv', encoding='utf-8-sig')
    income_annual = pd.read_csv(f'data/{stock_code}_income_statement_annual_ttm.csv', encoding='utf-8-sig')
    cashflow_annual = pd.read_csv(f'data/{stock_code}_cashflow_statement_annual_ttm.csv', encoding='utf-8-sig')
    
    # 生成报告
    generator = SummaryExcelGenerator(ts_token)
    output_path = f'data/{stock_code}_简要经营数据汇总.xlsx'
    generator.generate_report(
        balance_annual, income_annual, cashflow_annual,
        balance_annual, income_annual, cashflow_annual,
        stock_code, output_path
    )
