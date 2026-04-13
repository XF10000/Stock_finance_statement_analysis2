"""
Excel格式化输出模块
用于生成格式化的财务报表Excel文件
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import List, Optional


# 比率类型项目的关键词列表（这些项目的值将显示为百分比格式）
RATIO_KEYWORDS = [
    '率', '比例', '占比',
    '营业成本率',
    '毛利率',
    '销售费用率',
    '管理费用率',
    '研发费用率',
    '资产减值损失率',
    '营业外收支及其他占营业收入的比例',
    '息税前经营利润率',
    '实际所得税税率',
    '口径一收入现金含量',
    '成本费用付现率',
    '净利润现金含量',
    '非付现成本费用比经营活动产生的现金流量净额',
    '长期经营资产扩张性资本支出比例',
    '扩张性资本支出占长期资产期初净额的比例',
    '现金含量',
    '付现率',
    '营业收入占比',
    '费用率',
    '损失率',
    '税率',
    '利润率',
    '周转率',
]


def is_ratio_item(item_name: str) -> bool:
    """
    判断项目名称是否为比率类型
    
    Args:
        item_name: 项目名称
        
    Returns:
        是否为比率类型
    """
    if not item_name:
        return False
    item_name = str(item_name)
    return any(keyword in item_name for keyword in RATIO_KEYWORDS)


def save_formatted_balance_sheet(df: pd.DataFrame, filename: str, 
                                 highlight_keywords: Optional[List[str]] = None):
    """
    保存格式化的资产负债表到Excel
    
    Args:
        df: 资产负债表DataFrame
        filename: 输出文件名
        highlight_keywords: 需要高亮的关键词列表（如"小计"、"合计"等）
    """
    # 定义三级分类项目（用于颜色高亮）
    # 一级分类（最高层级）- 亮橙色
    level_1_items = ['资产总额', '资本总额']
    
    # 二级分类（中间层级）- 亮黄色
    level_2_items = ['金融资产合计', '长期股权投资', '经营资产合计', '有息债务合计', '所有者权益合计']
    
    # 三级分类（细分层级）- 浅黄色
    level_3_items = [
        '周转性经营投入合计', '营运资产小计', '营运负债小计',
        '长期经营资产合计', '短期债务', '长期债务', '归属于母公司股东权益合计'
    ]
    
    if highlight_keywords is None:
        highlight_keywords = ['小计', '合计', '资产总计', '负债总计', '股东权益合计', 
                             '流动资产合计', '非流动资产合计', '流动负债合计', '非流动负债合计',
                             '资产总额', '资本总额']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "资产负债表"
    
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        # 获取第一列的值（项目名称）
        item_name = row[0] if len(row) > 0 else None
        is_ratio = is_ratio_item(item_name) if r_idx > 1 else False
        
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 设置对齐方式
            if c_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # 表头格式
            if r_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # 分层级高亮显示
                if item_name:
                    item_name_str = str(item_name)
                    
                    # 一级分类 - 亮橙色 (RGB: 255, 153, 51)
                    if item_name_str in level_1_items:
                        cell.font = Font(bold=True, size=11, color='000000')
                        cell.fill = PatternFill(start_color='FF9933', end_color='FF9933', fill_type='solid')
                    
                    # 二级分类 - 亮黄色 (RGB: 255, 215, 0)
                    elif item_name_str in level_2_items:
                        cell.font = Font(bold=True, size=10, color='000000')
                        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                    
                    # 三级分类 - 浅黄色 (RGB: 255, 248, 220)
                    elif item_name_str in level_3_items:
                        cell.font = Font(bold=True, size=10, color='000000')
                        cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
                
                # 数字格式
                if c_idx > 1 and isinstance(value, (int, float)):
                    if is_ratio:
                        # 比率类型：数值为小数（如0.7567表示75.67%），使用Excel百分比格式
                        cell.number_format = '0.0%'
                    else:
                        # 金额类型：千分位分隔，无小数
                        cell.number_format = '#,##0'
    
    # 设置列宽
    ws.column_dimensions['A'].width = 35
    for col in range(2, len(df.columns) + 2):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    wb.save(filename)


def save_formatted_income_statement(df: pd.DataFrame, filename: str,
                                   highlight_keywords: Optional[List[str]] = None):
    """
    保存格式化的利润表到Excel，使用隔行变色和分类高亮提高可读性
    
    Args:
        df: 利润表DataFrame
        filename: 输出文件名
        highlight_keywords: 需要高亮的关键词列表
    """
    # 定义利润表的分类项目（用于颜色高亮）
    # 一级分类 - 亮橙色
    level_1_items = ['股权价值增加', '综合收益总额']
    
    # 二级分类 - 亮黄色
    level_2_items = [
        '息前税后经营利润', '税后利息费用', '净利润',
        '营业利润', '利润总额', '归属于母公司所有者的净利润'
    ]
    
    # 三级分类 - 浅黄色
    level_3_items = [
        '营业收入', '营业总成本', '息税前经营利润', '所得税费用',
        '其他综合收益', '归属于母公司所有者的综合收益总额'
    ]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "利润表"
    
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        item_name = row[0] if len(row) > 0 else None
        is_ratio = is_ratio_item(item_name) if r_idx > 1 else False
        
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 设置对齐方式
            if c_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # 表头格式
            if r_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # 分层级高亮显示
                is_highlighted = False
                if item_name:
                    item_name_str = str(item_name)
                    
                    # 一级分类 - 亮橙色
                    if item_name_str in level_1_items:
                        cell.font = Font(bold=True, size=11, color='000000')
                        cell.fill = PatternFill(start_color='FF9933', end_color='FF9933', fill_type='solid')
                        is_highlighted = True
                    
                    # 二级分类 - 亮黄色
                    elif item_name_str in level_2_items:
                        cell.font = Font(bold=True, size=10, color='000000')
                        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        is_highlighted = True
                    
                    # 三级分类 - 浅黄色
                    elif item_name_str in level_3_items:
                        cell.font = Font(bold=True, size=10, color='000000')
                        cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
                        is_highlighted = True
                
                # 如果不是高亮行，则使用隔行变色（斑马纹）
                if not is_highlighted and r_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                
                # 数字格式
                if c_idx > 1 and isinstance(value, (int, float)):
                    if is_ratio:
                        # 比率类型：数值为小数（如0.7567表示75.67%），使用Excel百分比格式
                        cell.number_format = '0.0%'
                    else:
                        cell.number_format = '#,##0'
    
    # 设置列宽
    ws.column_dimensions['A'].width = 40
    for col in range(2, len(df.columns) + 2):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    wb.save(filename)


def save_formatted_cashflow_statement(df: pd.DataFrame, filename: str,
                                     highlight_keywords: Optional[List[str]] = None):
    """
    保存格式化的现金流量表到Excel，使用隔行变色和分类高亮提高可读性
    
    Args:
        df: 现金流量表DataFrame
        filename: 输出文件名
        highlight_keywords: 需要高亮的关键词列表
    """
    # 定义现金流量表的分类项目（用于颜色高亮）
    # 一级分类 - 亮橙色
    level_1_items = ['自由现金流', '现金及现金等价物净增加额']
    
    # 二级分类 - 亮黄色
    level_2_items = [
        '经营活动产生的现金流量净额', '投资活动产生的现金流量净额',
        '筹资活动产生的现金流量净额', '扩张性资本支出'
    ]
    
    # 三级分类 - 浅黄色
    level_3_items = [
        '销售商品、提供劳务收到的现金', '购买商品、接受劳务支付的现金',
        '支付给职工以及为职工支付的现金', '支付的各项税费',
        '购建固定资产、无形资产和其他长期资产支付的现金'
    ]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "现金流量表"
    
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        item_name = row[0] if len(row) > 0 else None
        is_ratio = is_ratio_item(item_name) if r_idx > 1 else False
        
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 设置对齐方式
            if c_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # 表头格式
            if r_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # 分层级高亮显示
                is_highlighted = False
                if item_name:
                    item_name_str = str(item_name)
                    
                    # 一级分类 - 亮橙色
                    if item_name_str in level_1_items:
                        cell.font = Font(bold=True, size=11, color='000000')
                        cell.fill = PatternFill(start_color='FF9933', end_color='FF9933', fill_type='solid')
                        is_highlighted = True
                    
                    # 二级分类 - 亮黄色
                    elif item_name_str in level_2_items:
                        cell.font = Font(bold=True, size=10, color='000000')
                        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        is_highlighted = True
                    
                    # 三级分类 - 浅黄色
                    elif item_name_str in level_3_items:
                        cell.font = Font(bold=True, size=10, color='000000')
                        cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
                        is_highlighted = True
                
                # 如果不是高亮行，则使用隔行变色（斑马纹）
                if not is_highlighted and r_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                
                # 数字格式
                if c_idx > 1 and isinstance(value, (int, float)):
                    if is_ratio:
                        # 比率类型：数值为小数（如0.7567表示75.67%），使用Excel百分比格式
                        cell.number_format = '0.0%'
                    else:
                        cell.number_format = '#,##0'
    
    # 设置列宽
    ws.column_dimensions['A'].width = 45
    for col in range(2, len(df.columns) + 2):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    wb.save(filename)
