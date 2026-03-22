"""
带样式的财务报表导出模块
支持高亮、加粗、千分位数字格式
"""

import pandas as pd
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# 需要高亮的项目（总计、小计类）
HIGHLIGHT_ITEMS = [
    '金融资产合计',
    '长期股权投资',
    '经营资产合计',
    '周转性经营投入合计',
    '营运资产小计',
    '营运负债小计',
    '长期经营资产合计',
    '资产总额',
    '有息债务合计',
    '短期债务',
    '长期债务',
    '所有者权益合计',
    '归属于母公司股东权益合计',
    '少数股东权益',
    '资本总额',
]

# 需要加粗的项目（一级分类）
BOLD_ITEMS = [
    '金融资产合计',
    '长期股权投资',
    '经营资产合计',
    '资产总额',
    '有息债务合计',
    '所有者权益合计',
    '资本总额',
]


def format_number_thousands(value) -> str:
    """
    将数字格式化为千分位字符串，不带小数
    
    Args:
        value: 数值
        
    Returns:
        格式化后的字符串
    """
    if pd.isna(value) or value is None:
        return ''
    try:
        num = int(float(value))
        return f'{num:,}'
    except (ValueError, TypeError):
        return str(value)


def save_balance_sheet_to_excel_styled(
    df: pd.DataFrame,
    filepath: str,
    sheet_name: str = '资产负债表'
):
    """
    保存资产负债表到 Excel，带样式（高亮、加粗、千分位）
    
    Args:
        df: 重构后的资产负债表（第一列为'项目'，其余为日期列）
        filepath: 输出文件路径
        sheet_name: 工作表名称
    """
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 定义样式
    # 高亮填充（浅黄色）
    highlight_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    # 更亮的高亮（用于更重要的合计）
    strong_highlight_fill = PatternFill(start_color='FFE599', end_color='FFE599', fill_type='solid')
    
    # 字体样式
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    header_font = Font(bold=True, size=11)
    
    # 对齐方式
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # 边框
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # 写入表头
    headers = df.columns.tolist()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        # 表头背景色（浅灰色）
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # 写入数据行
    for row_idx, (idx, row) in enumerate(df.iterrows(), 2):
        item_name = row.iloc[0] if len(row) > 0 else ''
        
        # 判断是否需要高亮和加粗
        is_highlight = item_name in HIGHLIGHT_ITEMS
        is_bold = item_name in BOLD_ITEMS
        is_strong_highlight = item_name in ['资产总额', '资本总额', '所有者权益合计']
        
        # 选择填充色
        if is_strong_highlight:
            row_fill = strong_highlight_fill
        elif is_highlight:
            row_fill = highlight_fill
        else:
            row_fill = None
        
        # 选择字体
        if is_bold:
            row_font = bold_font
        else:
            row_font = normal_font
        
        # 写入每一列
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # 格式化值
            if col_idx == 1:
                # 第一列是项目名称
                cell.value = value
                cell.alignment = left_align
            else:
                # 数值列：格式化为千分位
                if pd.notna(value) and value is not None:
                    try:
                        # 保存原始数值用于计算，但显示格式化的字符串
                        num_val = float(value)
                        cell.value = int(num_val)
                        # 设置数字格式为千分位，无小数
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    cell.value = None
                cell.alignment = right_align
            
            # 应用样式
            if row_fill:
                cell.fill = row_fill
            cell.font = row_font
            cell.border = thin_border
    
    # 调整列宽
    # 第一列（项目名称）较宽
    ws.column_dimensions['A'].width = 25
    
    # 其他列根据内容自动调整，但设置最小宽度
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max(max_length + 2, 12), 20)
        ws.column_dimensions[column].width = adjusted_width
    
    # 冻结首行（固定表头）
    ws.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(filepath)


def save_to_excel_with_style(
    df: pd.DataFrame,
    filepath: str,
    sheet_name: str = 'Sheet1',
    highlight_items: Optional[List[str]] = None,
    bold_items: Optional[List[str]] = None
):
    """
    通用函数：保存 DataFrame 到带样式的 Excel
    
    Args:
        df: 数据DataFrame（第一列为项目名称）
        filepath: 输出文件路径
        sheet_name: 工作表名称
        highlight_items: 需要高亮的项目列表
        bold_items: 需要加粗的项目列表
    """
    if highlight_items is None:
        highlight_items = []
    if bold_items is None:
        bold_items = []
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 定义样式
    highlight_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # 写入表头
    headers = df.columns.tolist()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # 写入数据
    for row_idx, (idx, row) in enumerate(df.iterrows(), 2):
        item_name = row.iloc[0] if len(row) > 0 else ''
        
        is_highlight = item_name in highlight_items
        is_bold = item_name in bold_items
        
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if col_idx == 1:
                cell.value = value
                cell.alignment = left_align
            else:
                # 数值列格式化为千分位
                if pd.notna(value) and value is not None:
                    try:
                        num_val = float(value)
                        cell.value = int(num_val)
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    cell.value = None
                cell.alignment = right_align
            
            if is_highlight:
                cell.fill = highlight_fill
            if is_bold:
                cell.font = bold_font
            else:
                cell.font = normal_font
            cell.border = thin_border
    
    # 调整列宽
    ws.column_dimensions['A'].width = 25
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 12), 20)
        ws.column_dimensions[column].width = adjusted_width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存
    wb.save(filepath)
